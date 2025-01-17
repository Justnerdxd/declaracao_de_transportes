from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
import os


def gerar_excel_com_modelo(modelo_path, output_path, dados):
    """
    Gera um arquivo Excel baseado em um modelo e preenche os dados fornecidos.
    :param modelo_path: Caminho do modelo Excel.
    :param output_path: Caminho de saída do arquivo Excel gerado.
    :param dados: Dicionário com os dados a serem preenchidos.
    """
    try:
        # Carrega o modelo Excel
        workbook = load_workbook(modelo_path)
        sheet = workbook.active

        # Função auxiliar para definir valores no Excel
        def set_value(cell, value):
            """Define um valor em uma célula específica."""
            sheet[cell] = value

        # Preenche os campos no modelo Excel
        set_value("V5", dados.get("Codigo_Cliente", "N/A"))
        set_value("V6", dados.get("Codigo_Transportadora", "N/A"))
        set_value("D9", dados.get("Motorista", "N/A"))
        set_value("U9", dados.get("Telefone", "N/A"))
        set_value("N9", dados.get("CPF", "N/A"))
        set_value("V7", dados.get("Placa", "N/A"))
        set_value("E5", dados.get("Cliente_Nome", "N/A"))
        set_value("D6", dados.get("Cliente_Cidade", "N/A"))
        set_value("E7", dados.get("Transportadora_Nome", "N/A"))
        set_value("D8", dados.get("Transportadora_CNPJ", "N/A"))

        # Preenche as notas fiscais
        linha_inicial = 13  # Define a linha inicial no Excel para as notas fiscais
        for nf in dados.get("Notas_Fiscais", []):
            sheet[f"B{linha_inicial}"] = nf.get("Numero", "N/A")
            sheet[f"J{linha_inicial}"] = nf.get("Peso", "N/A")
            sheet[f"N{linha_inicial}"] = nf.get("Volumes", "N/A")
            linha_inicial += 1  # Incrementa a linha para a próxima nota fiscal

        # Salva o arquivo Excel gerado
        workbook.save(output_path)
        print(f"Arquivo Excel gerado com sucesso em: {output_path}")
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        raise


def transformar_excel_em_pdf(excel_path, pdf_path):
    """
    Converte um arquivo Excel gerado em um arquivo PDF.
    :param excel_path: Caminho do arquivo Excel.
    :param pdf_path: Caminho onde o PDF será salvo.
    """
    try:
        # Carrega o Excel
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Cria o PDF
        c = canvas.Canvas(pdf_path)
        c.setFont("Helvetica", 10)

        y = 800  # Posição inicial no PDF
        for row in sheet.iter_rows(values_only=True):
            linha = " | ".join(str(cell) for cell in row if cell is not None)
            c.drawString(30, y, linha)
            y -= 20

        c.save()
        print(f"PDF gerado em: {pdf_path}")
    except Exception as e:
        print(f"Erro ao converter Excel para PDF: {e}")
